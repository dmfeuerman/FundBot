import pandas as pd
import os
from csv import *
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side
import openpyxl.worksheet.worksheet


def Get_Cells(path, sb):
    wb = load_workbook(path, data_only=True)
    ws = wb["Sheet1"]
    df = pd.read_excel(path)

    directory = 'ProgramFiles/Csv_files/style_data_csv/' + sb + "/sorted"
    # print(directory)
    for filename in os.listdir(directory):
        # print(directory+ "/"+ filename)
        f = open(directory + "/" + filename, 'r')
        csv_reader = reader(f)
        bw_num = filename.strip(".csv").split("_")
        for line in csv_reader:
            line_name = line[0]
            # print(line_name + "line_name")
            for val in range(len(df)):
                name = df['Name'][val]
                # print(name)
                if line_name == name:

                    cell_val = str(colnum_string(int(bw_num[1]) + 1) + str(val + 2))
                    # print(cell_val)
                    try:
                        cell = ws[cell_val]
                        if bw_num[0] == "Best":
                            cell.font = Font(bold=True, size=22)
                            cell.fill = PatternFill("solid", fgColor="00ff00")
                        else:
                            cell.font = Font(italic=True, size=22)
                            cell.fill = PatternFill("solid", fgColor="ff0000")
                    except ValueError:
                        print(cell_val)

    wb.save(filename=path)


def colnum_string(n):
    string = ''
    while n > 0:
        n, remainder = divmod(n - 1, 26)
        string = chr(65 + remainder) + string
    return string


def Shade_NA(path, width):
    # load excel with its path
    # load excel with its path
    workbook = openpyxl.load_workbook(path)

    # print(workbook.get_sheet_names())
    worksheet = workbook['Sheet1']
    # print(worksheet.max_column)
    c = 2
    thin_border = Border(left=Side(style='thin', color="000000"),
                         right=Side(style='thin', color="000000"),
                         top=Side(style='thin', color="000000"),
                         bottom=Side(style='thin', color="000000"))
    for i in worksheet:
        for j in range(1, worksheet.max_column + 1):
            d = worksheet.cell(row=c, column=j)

            cell_val = str(colnum_string(j) + str(c))
            worksheet.column_dimensions[colnum_string(j)].width = width
            # print(cell_val)

            try:
                if d.value.strip() == "NA" or d.value.strip() == "N/A":
                    cell = worksheet[cell_val]
                    cell.fill = PatternFill("solid", fgColor="000000")



            except (ValueError, AttributeError):
                continue
            try:
                worksheet[cell_val] = float(d.value)
            except (ValueError, AttributeError):
                x = 0

        c = c + 1

    workbook.save(filename=path)


def insert_Columns(path, divider):
    workbook = openpyxl.load_workbook(path)

    # print(workbook.get_sheet_names())
    worksheet = workbook['Sheet1']
    columns = [5, 10, 15, 28, 31, 35, 39]
    for val in columns:
        worksheet.insert_cols(val)

    for j in columns:
        for c in range(1, len(worksheet['A']) + 1):
            d = worksheet.cell(row=j, column=j)

            cell_val = str(colnum_string(j) + str(c))
            # print(cell_val)

            try:
                # print(cell_val)
                cell = worksheet[cell_val]
                cell.fill = PatternFill("solid", fgColor="000000")

            except (ValueError, AttributeError):
                continue
        worksheet.column_dimensions[colnum_string(j)].width = divider
    workbook.save(filename=path)


def color(path):
    workbook = openpyxl.load_workbook(path)

    # print(workbook.get_sheet_names())
    worksheet = workbook['Sheet1']

    for j in range(9, 25):
        for c in range(2, len(worksheet['A']) + 1):
            d = worksheet.cell(row=c, column=j)

            cell_val = str(colnum_string(j) + str(c))

            cell = worksheet[cell_val]

            try:
                return_val = float(d.value.rstrip("%"))

                if return_val > 0:
                    cell.fill = PatternFill("solid", fgColor="00ff00")
                elif return_val <= 0:
                    cell.fill = PatternFill("solid", fgColor="ff0000")
            except (ValueError, AttributeError):
                continue

    workbook.save(filename=path)


def Highlight_Blue(path, fund_type):
    workbook = openpyxl.load_workbook(path)

    # print(workbook.get_sheet_names())
    worksheet = workbook['Sheet1']
    column = 4
    foregin = {33, 34, 36, 37, 38}
    us = {33, 35, 36, 37, 38}
    bonds = {33, 34, 35, 38}
    f_bond = {33, 34, 35, 38}

    for c in range(2, len(worksheet['D']) + 1):
        d = worksheet.cell(row=c, column=column)

        Style_blue(bonds, c, d, foregin, fund_type, us, f_bond, worksheet)
    workbook.save(filename=path)


def Style_blue(bonds, c, d, foregin, fund_type, us, f_bond, worksheet):
    if fund_type == "stock":
        if "Foreign" in d.value:
            for val in foregin:
                try:
                    type_val = worksheet.cell(row=c, column=val).value
                    if float(type_val) >= 7:
                        cell_val = str(colnum_string(val) + str(c))
                        cell = worksheet[cell_val]
                        cell.fill = PatternFill("solid", fgColor="add8e6")
                except ValueError:
                    continue
        else:
            for val in us:
                try:
                    type_val = worksheet.cell(row=c, column=val).value
                    if float(type_val) >= 7:
                        cell_val = str(colnum_string(val) + str(c))
                        cell = worksheet[cell_val]
                        cell.fill = PatternFill("solid", fgColor="add8e6")
                except ValueError:
                    continue
    else:
        if "Foreign" in d.value:
            for val in f_bond:
                try:
                    type_val = worksheet.cell(row=c, column=val).value
                    if float(type_val) >= 7:
                        cell_val = str(colnum_string(val) + str(c))
                        cell = worksheet[cell_val]
                        cell.fill = PatternFill("solid", fgColor="add8e6")
                except ValueError:
                    continue
        else:
            for val in bonds:
                try:
                    type_val = worksheet.cell(row=c, column=val).value
                    if float(type_val) >= 7:
                        cell_val = str(colnum_string(val) + str(c))
                        cell = worksheet[cell_val]
                        cell.fill = PatternFill("solid", fgColor="add8e6")
                except ValueError:
                    continue


def Divider_width(divider):
    f_2 = open("width_dividers.txt", 'r')
    val = f_2.readline()
    try:
        divider_new = float(val.strip())
    except ValueError:
        divider_new = divider
    if divider <= 0:
        divider_new = divider
    return divider_new, f_2


def Get_Cell_Width(width):
    f = open("width_cells.txt", 'r')
    val = f.readline()
    try:
        width_new = float(val.strip())
    except ValueError:
        width_new = width
    if width <= 0:
        width_new = width
    return f, width_new

def Fill_Borders(path):
    thin = Side(border_style="thin", color="000000")
    workbook = openpyxl.load_workbook(path)

    # print(workbook.get_sheet_names())
    worksheet = workbook['Sheet1']
    columns = worksheet.max_column

    for j in range(1, columns+1):
        for c in range(1, len(worksheet['A']) + 1):
            d = worksheet.cell(row=c, column=j)

            cell_val = str(colnum_string(j) + str(c))

            cell = worksheet[cell_val]

            try:
                cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
            except (ValueError, AttributeError):
                continue

    workbook.save(filename=path)

def Stylize_Data_Main():
    path1 = "Stock_mutual_funds.xlsx"
    path2 = "Bond_Mutual_funds.xlsx"
    sb1 = 'stocks'
    sb2 = 'bonds'
    width = 15
    divider = 5
    f, width = Get_Cell_Width(width)
    divider, f_2 = Divider_width(divider)

    Get_Cells(path1, sb1)
    color(path1)
    Shade_NA(path1, width)
    Highlight_Blue(path1, "stock")
    insert_Columns(path1, divider)
    Fill_Borders(path1)

    Get_Cells(path2, sb2)
    color(path2)
    Shade_NA(path2, width)
    Highlight_Blue(path2, "bond")
    insert_Columns(path2, divider)
    Fill_Borders(path2)
    f.close()
    f_2.close()
